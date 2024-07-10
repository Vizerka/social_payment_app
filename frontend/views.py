from django.shortcuts import render, get_object_or_404, redirect # type: ignore
from .models import Beneficiary, PaymentList, PaymentHistory, PaymentListBeneficiary, Application, Benefit, ApplicationList
from .forms import BeneficiaryForm, PaymentListForm, NewBeneficiaryForm, PaymentListBeneficiaryForm, UploadFileForm, ApplicationForm, PaymentListBulkEditForm, PaymentListBeneficiaryFormSet
from django.db.models import Q # type: ignore
from django.core.paginator import Paginator  # type: ignore # Import Paginator
from django.contrib.auth.decorators import login_required # type: ignore
from django.forms import inlineformset_factory # type: ignore
from django.core.exceptions import ValidationError # type: ignore
import openpyxl # type: ignore
from openpyxl.utils import get_column_letter # type: ignore
from django.http import HttpResponse # type: ignore
from io import BytesIO
from django.contrib import messages # type: ignore
from django.db import transaction, IntegrityError # type: ignore
from .tables import PaymentListBeneficiaryTable

def index(request): #zdefiniowanie strony index
    return render(request, 'frontend/index.html')

@login_required
def beneficiary_list(request):
    query = request.GET.get('q')
    if query:
        beneficiaries = Beneficiary.objects.filter(
            Q(first_name__icontains=query) | 
            Q(last_name__icontains=query) | 
            Q(place__icontains=query)
        )
    else:
        beneficiaries = Beneficiary.objects.all()

    active_count = Beneficiary.objects.filter(is_alive=True).count()
    inactive_count = Beneficiary.objects.filter(is_alive=False).count()
    
    paginator = Paginator(beneficiaries, 50)  # 50 beneficjentów na stronę
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'frontend/beneficiary_list.html', {'page_obj': page_obj, 'active_count':active_count, 'inactive_count':inactive_count})

@login_required
def beneficiary_add(request): #zdefiniowanie manualnego dodawania użytkowników 
    if request.method != 'POST':
        form = NewBeneficiaryForm()
    else:
        form = NewBeneficiaryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('frontend:beneficiary_add')
    return render(request, 'frontend/beneficiary_add.html', {'form':form})

@login_required
def beneficiary_detail(request, pk):
    beneficiary = get_object_or_404(Beneficiary, pk=pk)
    payment_history = PaymentHistory.objects.filter(beneficiary=beneficiary)
    return render(request, 'frontend/beneficiary_detail.html', {'beneficiary': beneficiary, 'payment_history': payment_history})

@login_required
def payment_add(request):
    if request.method == 'POST':
        form = PaymentListForm(request.POST)
        if form.is_valid():
            payment_list = form.save()
            return redirect('frontend:payment_add_beneficiary', pk=payment_list.pk)
    else:
        form = PaymentListForm()
    return render(request, 'frontend/payment_add.html', {'form': form})

def payment_add_beneficiary(request, pk):
    payment_list = get_object_or_404(PaymentList, pk=pk)
    
    if request.method == 'POST':
        form = PaymentListBeneficiaryForm(request.POST)
        if form.is_valid():
            payment_beneficiary = form.save(commit=False)
            payment_beneficiary.payment_list = payment_list
            
            beneficiary = payment_beneficiary.beneficiary
            
            # Sprawdzenie, czy beneficjent jest żywy
            if not beneficiary.is_alive:
                form.add_error('beneficiary', 'Ten beneficjent nie żyje i nie może być dodany do listy wypłat.')
            elif PaymentListBeneficiary.objects.filter(payment_list=payment_list, beneficiary=beneficiary).exists():
                form.add_error('beneficiary', 'Ten beneficjent już istnieje na tej liście wypłat.')
            else:
                payment_beneficiary.save()
                PaymentHistory.objects.create(
                    beneficiary=beneficiary,
                    payment_list=payment_list,
                    amount=payment_beneficiary.amount
                )
                messages.success(request, f'Beneficjent {beneficiary} został dodany do listy wypłat.')
                return redirect('frontend:payment_add_beneficiary', pk=payment_list.pk)
    else:
        form = PaymentListBeneficiaryForm()
    
    return render(request, 'frontend/payment_add_beneficiary.html', {'form': form, 'payment_list': payment_list})

@login_required
def payment_detail(request, pk):
    payment = get_object_or_404(PaymentList, pk=pk)
    payment_list_beneficiaries = PaymentListBeneficiary.objects.filter(payment_list=payment)
    
    beneficiaries = [
        {
            "first_name": plb.beneficiary.first_name,
            "last_name": plb.beneficiary.last_name,
            "place": plb.beneficiary.place,
            "bank_account_number": plb.beneficiary.bank_account_number,
            "amount": plb.amount
        }
        for plb in payment_list_beneficiaries
    ]
    
    return render(request, 'frontend/payment_detail.html', {'payment': payment, 'beneficiaries': beneficiaries})

@login_required
def payment_list(request):
    payments = PaymentList.objects.all()
    return render(request, 'frontend/payment_list.html', {'payments': payments})

@login_required
def export_payment_list_to_excel(request, pk):
    payment_list = get_object_or_404(PaymentList, pk=pk)
    payment_list_beneficiaries = PaymentListBeneficiary.objects.filter(payment_list=payment_list)

    # Create a workbook and a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = payment_list.name

    # Define the titles for columns
    columns = ['rachunek obcy', 'nr konta', 'imie i nazwisko', 'kwota', 'Tytuł', 'data']
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Fill the worksheet with data
    for plb in payment_list_beneficiaries:
        row_num += 1
        worksheet.cell(row=row_num, column=1, value=f"9512401037111001109117246")
        worksheet.cell(row=row_num, column=3, value=f"{plb.beneficiary.first_name} {plb.beneficiary.last_name}")
        worksheet.cell(row=row_num, column=2, value=plb.beneficiary.bank_account_number)
        worksheet.cell(row=row_num, column=5, value=f"{plb.beneficiary.place} {payment_list.name}")
        worksheet.cell(row=row_num, column=4, value=plb.amount)
        worksheet.cell(row=row_num, column=6, value=payment_list.date_added)

    # Adjust column widths
    for col_num, column_title in enumerate(columns, 1):
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = 20

    # Save the workbook to a bytes buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    # Create an HTTP response with Excel content
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename={payment_list.name}.xlsx'

    return response

@login_required
def import_beneficiaries(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            wb = openpyxl.load_workbook(file)
            sheet = wb.active
            skipped_beneficiaries = []

            for row in sheet.iter_rows(min_row=2, values_only=True):
                first_name, last_name, place, bank_account_number = row
                if Beneficiary.objects.filter(first_name=first_name, last_name=last_name, place=place).exists():
                    skipped_beneficiaries.append(f'{first_name} {last_name} ({place})')
                    continue

                Beneficiary.objects.create(
                    first_name=first_name,
                    last_name=last_name,
                    place=place,
                    bank_account_number=bank_account_number,)

            if skipped_beneficiaries:
                messages.warning(request, f'Import przebiegł pomyślnie, pominięci zostali następujący beneficjenci: {", ".join(skipped_beneficiaries)}')
            else:
                messages.success(request, 'Pomyślnie zaimportowano wszystkich beneficjentów')

            return redirect('frontend:beneficiary_list')
    else:
        form = UploadFileForm()
    return render(request, 'frontend/import_beneficiaries.html', {'form': form})

@login_required
def beneficiary_edit(request, pk):
    beneficiary = get_object_or_404(Beneficiary, pk=pk)
    
    if request.method == 'POST':
        form = BeneficiaryForm(request.POST, instance=beneficiary)
        if form.is_valid():
            with transaction.atomic():
                if Beneficiary.objects.filter(pk=pk, version=form.cleaned_data['version']).exists():
                    beneficiary = form.save(commit=False)
                    beneficiary.version += 1
                    beneficiary.save()
                    messages.success(request, 'Beneficjent został pomyślnie zaktualizowany.')
                    return redirect('frontend:beneficiary_detail', pk=beneficiary.pk)
                else:
                    form.add_error(None, 'Ten rekord został już zaktualizowany przez innego użytkownika.')
    else:
        form = BeneficiaryForm(instance=beneficiary)
        form.fields['version'].initial = beneficiary.version
    
    return render(request, 'frontend/beneficiary_edit.html', {'form': form, 'beneficiary': beneficiary})

def beneficiary_list_full(request):
    query = request.GET.get('q')
    if query:
        beneficiaries = Beneficiary.objects.filter(
            Q(first_name__icontains=query) | 
            Q(last_name__icontains=query) | 
            Q(place__icontains=query)
        )
    else:
        beneficiaries = Beneficiary.objects.all()

    active_count = Beneficiary.objects.filter(is_alive=True).count()
    inactive_count = Beneficiary.objects.filter(is_alive=False).count()
    
    paginator = Paginator(beneficiaries, 50)  # 50 beneficjentów na stronę
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'frontend/beneficiary_list_full.html', {'page_obj': page_obj, 'active_count':active_count, 'inactive_count':inactive_count})

@login_required
def application_create(request):
    if request.method == 'POST':
        form = ApplicationForm(request.POST)
        if form.is_valid():
            application = form.save(commit=False)
            application.save()
            messages.success(request, 'Application created successfully')
            return redirect('frontend:application_list_all')
    else:
        form = ApplicationForm()
    return render(request, 'frontend/application_form.html', {'form': form})

@login_required
def application_list_all(request):
    applications = Application.objects.all()
    application_lists = {}
    for application in applications:
        key = f"{application.benefit.name} {application.date_submitted.year}"
        if key not in application_lists:
            application_lists[key] = {
                'benefit': application.benefit,
                'year': application.date_submitted.year,
                'applications': []
            }
        application_lists[key]['applications'].append(application)
    return render(request, 'frontend/application_list_all.html', {'application_lists': application_lists.values()})

@login_required
def application_list_detail(request, benefit_id, year):
    benefit = get_object_or_404(Benefit, id=benefit_id)
    applications = Application.objects.filter(benefit=benefit, date_submitted__year=year)
    if request.method == 'POST':
        selected_applications = request.POST.getlist('applications')
        payment_list = PaymentList.objects.create(name=f"{benefit.name} {year}")
        for app_id in selected_applications:
            application = get_object_or_404(Application, id=app_id)
            PaymentListBeneficiary.objects.create(
                payment_list=payment_list,
                beneficiary=application.beneficiary,
                amount=application.amount
            )
            PaymentHistory.objects.create(
                beneficiary=application.beneficiary,
                payment_list=payment_list,
                amount=application.amount
            )
        messages.success(request, 'Lista wypłat utworzona poprawnie!')
        return redirect('frontend:payment_detail', pk=payment_list.pk)
    return render(request, 'frontend/application_list_detail.html', {'applications': applications, 'benefit': benefit, 'year': year})